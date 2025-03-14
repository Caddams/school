import os
import openpyxl
from tkinter import Tk, Label, Entry, Button, Listbox, messagebox, Toplevel
import tkinter as tk

# Ensure the Excel file exists
def ensure_excel_file_exists():
    file_name = "Charge Codes.xlsx"
    if not os.path.exists(file_name):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Charge Codes"
        ws["A1"] = "Entry Number"
        ws["B1"] = "Charge Code"
        wb.save(file_name)
    return file_name

# Add a new entry to the Excel file
def add_entry(charge_code):
    if len(charge_code) > 15:
        messagebox.showerror("Error", "Charge code cannot exceed 15 characters.")
        return

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    next_entry_number = ws.max_row  # Start numbering at 1, skipping the header
    ws.cell(row=next_entry_number + 1, column=1, value=next_entry_number)
    ws.cell(row=next_entry_number + 1, column=2, value=charge_code)
    wb.save(excel_file)

    update_listbox()

# Update the listbox to display all entry numbers and charge codes
def update_listbox():
    listbox.delete(0, "end")
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2, values_only=True):
        entry_number, charge_code = row
        listbox.insert("end", f"{entry_number}: {charge_code}")

# Open a window to edit a selected charge code
def on_listbox_select(event):
    selection = listbox.curselection()
    if not selection:
        return

    selected_item = listbox.get(selection[0])
    entry_number, current_code = selected_item.split(": ", 1)

    def save_edit():
        new_code = edit_entry.get().strip()
        if len(new_code) > 15:
            messagebox.showerror("Error", "Charge code cannot exceed 15 characters.")
            return

        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
            if row[0].value == int(entry_number):
                row[1].value = new_code
                break

        wb.save(excel_file)
        update_listbox()
        edit_window.destroy()

    edit_window = Toplevel(root)
    edit_window.title("Edit Charge Code")
    Label(edit_window, text=f"Editing Entry {entry_number}").pack(pady=5)
    edit_entry = Entry(edit_window, width=20)
    edit_entry.insert(0, current_code)
    edit_entry.pack(pady=5)
    Button(edit_window, text="Save", command=save_edit).pack(pady=5)

# Handle the add button click
def on_add_button_click():
    charge_code = entry.get().strip()
    if charge_code:
        add_entry(charge_code)
        entry.delete(0, "end")
    else:
        messagebox.showerror("Error", "Please enter a charge code.")

# Function to handle back button click (close the application)
def handle_back_button():
    root.destroy()

# Initialize the application
excel_file = ensure_excel_file_exists()

root = Tk()
root.title("Charge Code Manager")
root.state('zoomed')  # Set to windowed fullscreen mode

# Fonts
title_font = ("Helvetica", 24, "bold")

# Add the "Charge Code Management System" label at the top
title_label = tk.Label(root, text="Charge Code Management System", font=title_font, anchor="center")
title_label.pack(pady=(10, 20))

# Add the "Back" button at the top-right corner
back_button = tk.Button(root, text="Back", width=10, command=handle_back_button)
back_button.pack(anchor="ne", padx=10, pady=10)

Label(root, text="Enter Charge Code (Max 15 Characters):").pack(pady=5)

entry = Entry(root, width=20)
entry.pack(pady=5)

Button(root, text="Add Charge Code", command=on_add_button_click).pack(pady=5)

Label(root, text="Current Charge Codes:").pack(pady=5)

listbox = Listbox(root, width=50, height=10)
listbox.pack(pady=5)
listbox.bind("<<ListboxSelect>>", on_listbox_select)

update_listbox()

root.mainloop()
